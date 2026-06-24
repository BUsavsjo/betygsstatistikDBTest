from __future__ import annotations

from collections import Counter
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .constants import GradeSpec, NpSpec
from .io import read_text_rows
from .metrics import clean


SENSITIVE_COLUMNS = {"PersonNr", "Fornamn", "Efternamn"}


@dataclass(frozen=True)
class DocumentedField:
    position: int
    name: str
    normalized_name: str
    description: str
    allowed_chars: str
    max_length: str


@dataclass(frozen=True)
class ControlCase:
    sheet_name: str
    kind: str
    arskurs: int
    raw_dir: Path
    documentation_path: Path
    import_columns: list[str]


def normalize_documented_column(name: str) -> str:
    return {
        "M1 (språk)": "M1_sprak",
        "M1 (betyg)": "M1_betyg",
        "M2 (språk)": "M2_sprak",
        "M2 (betyg)": "M2_betyg",
        "ML (språk)": "ML_sprak",
        "ML (betyg)": "ML_betyg",
        "M1 (sprÃ¥k)": "M1_sprak",
        "M2 (sprÃ¥k)": "M2_sprak",
        "ML (sprÃ¥k)": "ML_sprak",
        "OVR": "Ovr",
    }.get(name, name)


def datafile_year_suffix(lasar: str) -> str:
    end_year = lasar.split("-")[-1]
    if len(end_year) == 4 and end_year.isdigit():
        return end_year
    return datetime.now().strftime("%Y")


def grade_documentation_path(documentation_base: Path, arskurs: int, lasar: str) -> Path:
    suffix = datafile_year_suffix(lasar)
    candidates = [
        documentation_base / f"datafilsbeskrivning_betyg_ak{arskurs}_{suffix}.xlsx",
        documentation_base / f"datafilsbeskrivning-betyg_ak{arskurs}_{suffix}.xlsx",
    ]
    for path in candidates:
        if path.exists():
            return path
    return candidates[0]


def np_documentation_path(documentation_base: Path, arskurs: int, lasar: str) -> Path:
    suffix = datafile_year_suffix(lasar)
    return documentation_base / "np" / f"datafilsbeskrivning_np{arskurs}_{suffix}.xlsx"


def read_documented_fields(path: Path) -> list[DocumentedField]:
    if not path.exists():
        return []

    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook.worksheets[0]
        header_row = None
        for row_number, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            first_cell = clean(row[0]) if row and row[0] is not None else ""
            if first_cell in {"Variabelnamn", "Variabel"}:
                header_row = row_number
                break
        if header_row is None:
            return []

        fields: list[DocumentedField] = []
        for row in sheet.iter_rows(min_row=header_row + 1, values_only=True):
            name = clean(row[0]) if row and row[0] is not None else ""
            description = clean(row[1]) if len(row) > 1 and row[1] is not None else ""
            if not name or not description:
                continue
            fields.append(DocumentedField(
                position=len(fields) + 1,
                name=name,
                normalized_name=normalize_documented_column(name),
                description=description,
                allowed_chars=clean(row[2]) if len(row) > 2 and row[2] is not None else "",
                max_length=clean(row[3]) if len(row) > 3 and row[3] is not None else "",
            ))
        return fields
    finally:
        workbook.close()


def read_raw_rows(raw_dir: Path) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    rows: list[dict[str, Any]] = []
    errors: list[dict[str, Any]] = []
    if not raw_dir.exists():
        errors.append({"fil": "", "rad": "", "typ": "MAPP_SAKNAS", "detalj": str(raw_dir)})
        return rows, errors

    for path in sorted(raw_dir.glob("*.txt")):
        try:
            for row_number, values in enumerate(read_text_rows(path), start=1):
                rows.append({"file": path.name, "row_number": row_number, "values": values})
        except UnicodeDecodeError as error:
            errors.append({"fil": path.name, "rad": "", "typ": "KAN_INTE_LASA_FIL", "detalj": str(error)})
    return rows, errors


def parse_max_length(value: str) -> int | None:
    text = clean(value)
    if not text:
        return None
    parts = [part for part in text.replace(" ", "").split("-") if part.isdigit()]
    if parts:
        return int(parts[-1])
    return int(text) if text.isdigit() else None


def example_value(column_name: str, value: str) -> str:
    if column_name in SENSITIVE_COLUMNS:
        return "[maskerat]"
    return value


def summarize_position(rows: list[dict[str, Any]], index: int, column_name: str) -> dict[str, Any]:
    values = [clean(row["values"][index]) if index < len(row["values"]) else "" for row in rows]
    filled = [value for value in values if value]
    examples: list[str] = []
    for value, _count in Counter(filled).most_common():
        display = example_value(column_name, value)
        if display not in examples:
            examples.append(display)
        if len(examples) == 5:
            break
    return {
        "Antal TXT-rader": len(rows),
        "Antal ifyllda": len(filled),
        "Antal tomma": len(values) - len(filled),
        "Max faktisk längd": max((len(value) for value in values), default=0),
        "Exempelvärden": ", ".join(examples),
    }


def status_for_row(
    documented: DocumentedField | None,
    import_column: str,
    summary: dict[str, Any],
) -> tuple[str, str]:
    if documented is None and import_column:
        return "SAKNAS_I_DATAFILSBESKRIVNING", "Importkoden förväntar sig ett fält som inte finns i datafilsbeskrivningen."
    if documented is not None and not import_column:
        return "SAKNAS_I_IMPORTKOD", "Datafilsbeskrivningen har ett fält som importkoden inte läser."
    if documented is not None and import_column and documented.normalized_name != import_column:
        return "NAMN_AVVIKER", "Fältet finns på samma position men namnet skiljer sig."

    max_length = parse_max_length(documented.max_length if documented else "")
    if max_length is not None and int(summary["Max faktisk längd"]) > max_length:
        return "ÖVER_MAXLÄNGD", "Minst ett TXT-värde är längre än datafilsbeskrivningens maxlängd."
    return "OK", ""


def build_case_rows(case: ControlCase) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    documented_fields = read_documented_fields(case.documentation_path)
    raw_rows, errors = read_raw_rows(case.raw_dir)
    max_columns = max(
        len(documented_fields),
        len(case.import_columns),
        max((len(row["values"]) for row in raw_rows), default=0),
    )

    rows: list[dict[str, Any]] = []
    for index in range(max_columns):
        documented = documented_fields[index] if index < len(documented_fields) else None
        import_column = case.import_columns[index] if index < len(case.import_columns) else ""
        compare_name = documented.normalized_name if documented else ""
        column_name = import_column or compare_name
        summary = summarize_position(raw_rows, index, column_name)
        status, comment = status_for_row(documented, import_column, summary)
        rows.append({
            "Position": index + 1,
            "Datafilsbeskrivning: Variabelnamn": documented.name if documented else "",
            "Datafilsbeskrivning: Normaliserat namn": compare_name,
            "Importkod: Kolumnnamn": import_column,
            "Matchar namn": "Ja" if documented and import_column and compare_name == import_column else "Nej",
            "Fältinnehåll": documented.description if documented else "",
            "Tillåtna tecken": documented.allowed_chars if documented else "",
            "Max längd": documented.max_length if documented else "",
            **summary,
            "Status": status,
            "Kommentar": comment,
        })

    expected_columns = len(case.import_columns)
    for raw_row in raw_rows:
        actual_columns = len(raw_row["values"])
        if actual_columns != expected_columns:
            errors.append({
                "fil": raw_row["file"],
                "rad": raw_row["row_number"],
                "typ": "FEL_KOLUMNANTAL",
                "detalj": f"Förväntat {expected_columns}, faktiskt {actual_columns}",
                "blad": case.sheet_name,
            })
    if not documented_fields:
        errors.append({
            "fil": case.documentation_path.name,
            "rad": "",
            "typ": "DATAFILSBESKRIVNING_SAKNAS",
            "detalj": str(case.documentation_path),
            "blad": case.sheet_name,
        })
    return rows, errors


def write_table(sheet: Any, rows: list[dict[str, Any]]) -> None:
    if not rows:
        return
    headers = list(rows[0])
    sheet.append(headers)
    for row in rows:
        sheet.append([row.get(header, "") for header in headers])

    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    widths = {
        "A": 10,
        "B": 28,
        "C": 28,
        "D": 24,
        "E": 14,
        "F": 48,
        "G": 18,
        "H": 12,
        "I": 14,
        "J": 14,
        "K": 12,
        "L": 16,
        "M": 42,
        "N": 24,
        "O": 58,
    }
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions


def write_readme(sheet: Any, lasar: str, output_path: Path, cases: list[ControlCase]) -> None:
    rows = [
        ("Kontrollfil", "Jämför SCB:s datafilsbeskrivningar mot importkodens fält och faktisk TXT-inläsning."),
        ("Läsår", lasar),
        ("Skapad", datetime.now().strftime("%Y-%m-%d %H:%M")),
        ("Fil", str(output_path)),
        ("Personuppgifter", "Exempelvärden för PersonNr, Förnamn och Efternamn maskas."),
        ("Git/publicering", "Filen skapas under data/output och ska inte publiceras eller läggas i Git."),
        ("", ""),
        ("Blad", "Datafilsbeskrivning"),
    ]
    for case in cases:
        rows.append((case.sheet_name, str(case.documentation_path)))
    for row in rows:
        sheet.append(row)
    sheet.column_dimensions["A"].width = 22
    sheet.column_dimensions["B"].width = 120
    for cell in sheet[1]:
        cell.font = Font(bold=True)


def create_control_cases(
    *,
    lasar: str,
    grade_raw: Path,
    np_raw: Path,
    documentation_base: Path,
    grade_specs: list[GradeSpec],
    np_specs: list[NpSpec],
) -> list[ControlCase]:
    cases: list[ControlCase] = []
    for spec in grade_specs:
        cases.append(ControlCase(
            sheet_name=f"betyg_ak{spec.arskurs}",
            kind="betyg",
            arskurs=spec.arskurs,
            raw_dir=grade_raw / lasar / spec.raw_folder,
            documentation_path=grade_documentation_path(documentation_base, spec.arskurs, lasar),
            import_columns=spec.columns,
        ))
    for spec in np_specs:
        cases.append(ControlCase(
            sheet_name=f"np_ak{spec.arskurs}",
            kind="np",
            arskurs=spec.arskurs,
            raw_dir=np_raw / lasar / spec.raw_folder,
            documentation_path=np_documentation_path(documentation_base, spec.arskurs, lasar),
            import_columns=spec.columns,
        ))
    return cases


def write_datafile_control_workbook(
    output_path: Path,
    *,
    lasar: str,
    cases: list[ControlCase],
) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    workbook.remove(workbook.active)

    readme = workbook.create_sheet("README")
    write_readme(readme, lasar, output_path, cases)

    all_errors: list[dict[str, Any]] = []
    for case in cases:
        sheet = workbook.create_sheet(case.sheet_name)
        rows, errors = build_case_rows(case)
        write_table(sheet, rows)
        all_errors.extend({**error, "blad": error.get("blad", case.sheet_name)} for error in errors)

    error_sheet = workbook.create_sheet("avvikelser")
    if all_errors:
        write_table(error_sheet, all_errors)
    else:
        write_table(error_sheet, [{"blad": "", "fil": "", "rad": "", "typ": "OK", "detalj": "Inga avvikelser hittades."}])

    for sheet in workbook.worksheets:
        for column_number in range(1, sheet.max_column + 1):
            letter = get_column_letter(column_number)
            if sheet.column_dimensions[letter].width is None:
                sheet.column_dimensions[letter].width = 18

    workbook.save(output_path)
