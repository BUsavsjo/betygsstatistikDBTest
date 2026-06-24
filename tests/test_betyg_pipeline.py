from __future__ import annotations

import sys
import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

sys.path.insert(0, str(Path(__file__).resolve().parents[1] / "src"))

from openpyxl import Workbook, load_workbook

from config_paths import DEFAULT_LASAR, resolve_paths
from betyg.constants import NP_SPECS, SPECS, AK9_SUBJECTS
from betyg.datafile_control import build_case_rows, ControlCase
from betyg.io import publish_processed_json, read_np_files
from betyg.metrics import eligibility, gender_from_personnr, kolada_grade6_all_subjects_percentage, merit, overview, sv_sva_group
from betyg.np_data import aggregate_np
from betyg.pipeline import control_rows, np_import_diagnostics


PROJECT_ROOT = Path(__file__).resolve().parents[1]


def documented_columns(path: Path) -> list[str]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.worksheets[0]
    header_row = None
    for row_number, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        first_cell = str(row[0]).strip() if row and row[0] is not None else ""
        if first_cell in {"Variabelnamn", "Variabel"}:
            header_row = row_number
            break
    if header_row is None:
        raise AssertionError(f"Saknar variabelrubrik i {path}")

    columns = []
    for row in sheet.iter_rows(min_row=header_row + 1, values_only=True):
        name = str(row[0]).strip() if row and row[0] is not None else ""
        description = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ""
        if name and description:
            columns.append(normalize_documented_column(name))
    return columns


def normalize_documented_column(name: str) -> str:
    return {
        "M1 (språk)": "M1_sprak",
        "M1 (betyg)": "M1_betyg",
        "M2 (språk)": "M2_sprak",
        "M2 (betyg)": "M2_betyg",
        "ML (språk)": "ML_sprak",
        "ML (betyg)": "ML_betyg",
        "OVR": "Ovr",
    }.get(name, name)


class DatafileDescriptionContractTests(unittest.TestCase):
    def test_grade_import_columns_match_2026_datafile_descriptions(self) -> None:
        cases = {
            6: PROJECT_ROOT / "data" / "dokumentation" / "datafilsbeskrivning_betyg_ak6_2026.xlsx",
            9: PROJECT_ROOT / "data" / "dokumentation" / "datafilsbeskrivning_betyg_ak9_2026.xlsx",
        }
        for arskurs, path in cases.items():
            with self.subTest(arskurs=arskurs):
                self.assertEqual(SPECS[arskurs].columns, documented_columns(path))

    def test_np_import_columns_match_2026_datafile_descriptions(self) -> None:
        cases = {
            3: PROJECT_ROOT / "data" / "dokumentation" / "np" / "datafilsbeskrivning_np3_2026.xlsx",
            6: PROJECT_ROOT / "data" / "dokumentation" / "np" / "datafilsbeskrivning_np6_2026.xlsx",
            9: PROJECT_ROOT / "data" / "dokumentation" / "np" / "datafilsbeskrivning_np9_2026.xlsx",
        }
        for arskurs, path in cases.items():
            with self.subTest(arskurs=arskurs):
                self.assertEqual(NP_SPECS[arskurs].columns, documented_columns(path))

    def test_datafile_control_compares_documentation_import_and_txt_values(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            root = Path(tmp_dir)
            docs = root / "datafilsbeskrivning.xlsx"
            raw_dir = root / "raw"
            raw_dir.mkdir()
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["Variabelnamn", "Fältinnehåll", "Tillåtna tecken", "(Max) längd"])
            sheet.append(["PersonNr", "Elevens personnummer", "BS", 12])
            sheet.append(["M1 (språk)", "Modernt språk 1", "A", 20])
            workbook.save(docs)
            (raw_dir / "test.txt").write_text("201001011234;SPA\n", encoding="utf-8")

            rows, errors = build_case_rows(ControlCase(
                sheet_name="test",
                kind="betyg",
                arskurs=6,
                raw_dir=raw_dir,
                documentation_path=docs,
                import_columns=["PersonNr", "M1_sprak"],
            ))

        self.assertEqual(errors, [])
        self.assertEqual(rows[0]["Status"], "OK")
        self.assertEqual(rows[0]["Exempelvärden"], "[maskerat]")
        self.assertEqual(rows[1]["Datafilsbeskrivning: Normaliserat namn"], "M1_sprak")
        self.assertEqual(rows[1]["Status"], "OK")

    def test_np_ak3_import_accepts_compact_31_column_format(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            root = Path(tmp_dir)
            raw_dir = root / "2025-2026" / "ak3"
            raw_dir.mkdir(parents=True)
            (raw_dir / "np3.txt").write_text(
                "Procapita;2026-06-24;14.0.0.00000;201601159666;13654995;3A;1;1;1;1;0;1;1;12;15;9;3;17;12;3A;2;1;1;1;1;1;1;1;1;15;16\n",
                encoding="utf-8",
            )

            rows, diagnostics = read_np_files(root, "2025-2026", NP_SPECS[3])

        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]["MA_G1_PR"], "")
        self.assertEqual(rows[0]["MA_G2_PR"], "")
        self.assertEqual(rows[0]["MA_G_POANG"], "12")
        self.assertEqual(rows[0]["SV_H_PR"], "1")
        self.assertFalse(any(item["message"] == "wrong_column_count" for item in diagnostics if item["level"] == "error"))


class MetricsTests(unittest.TestCase):
    def test_gender_from_personnr_uses_second_last_digit(self) -> None:
        self.assertEqual(gender_from_personnr("20100101-1234"), "Pojkar")
        self.assertEqual(gender_from_personnr("20100101-1244"), "Flickor")
        self.assertEqual(gender_from_personnr(""), "okänt")

    def test_sv_sva_group_prefers_distinct_states(self) -> None:
        self.assertEqual(sv_sva_group({"Sv": "C", "Sva": ""}), "SV")
        self.assertEqual(sv_sva_group({"Sv": "", "Sva": "B"}), "SVA")
        self.assertEqual(sv_sva_group({"Sv": "A", "Sva": "B"}), "SV_och_SVA")
        self.assertEqual(sv_sva_group({"Sv": "", "Sva": ""}), "oklar")

    def test_merit_uses_best_sv_sva_and_best_modern_language(self) -> None:
        row = {subject: "E" for subject in AK9_SUBJECTS if subject not in {"Sv", "Sva"}}
        row.update({"Sv": "C", "Sva": "A", "M1_betyg": "B", "M2_betyg": "D"})
        merit16, merit17 = merit(row, AK9_SUBJECTS)
        self.assertEqual(merit16, 170.0)
        self.assertEqual(merit17, 187.5)

    def test_eligibility_requires_core_subjects(self) -> None:
        row = {subject: "E" for subject in AK9_SUBJECTS}
        row["Sv"] = "F"
        row["Sva"] = ""
        result = eligibility(row, AK9_SUBJECTS)
        self.assertFalse(result["behorig_yrkesprogram"])
        self.assertFalse(result["behorig_hogskoleforberedande_nagot_program"])

    def test_kolada_grade6_all_subjects_percentage_uses_95_rule(self) -> None:
        self.assertEqual(kolada_grade6_all_subjects_percentage(39, 40), 95.0)
        self.assertEqual(kolada_grade6_all_subjects_percentage(35, 40), 87.5)


class PublishTests(unittest.TestCase):
    def test_publish_processed_json_copies_only_whitelisted_files(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            root = Path(tmp_dir)
            output_base = root / "output"
            processed_base = root / "processed"
            source_dir = output_base / "2025-2026" / "json"
            source_dir.mkdir(parents=True)
            (source_dir / "manifest.json").write_text("{}", encoding="utf-8")
            (source_dir / "betygsstatistik_oversikt.json").write_text("[]", encoding="utf-8")
            (source_dir / "secret.json").write_text("{}", encoding="utf-8")

            target_dir = publish_processed_json(output_base, processed_base, "2025-2026")

            self.assertTrue((target_dir / "manifest.json").exists())
            self.assertTrue((target_dir / "betygsstatistik_oversikt.json").exists())
            self.assertFalse((target_dir / "secret.json").exists())


class ControlRowsTests(unittest.TestCase):
    def test_overview_includes_gender_and_sv_sva_segments(self) -> None:
        rows = [
            {
                "Skolenhetskod": "123",
                "kon": "Pojkar",
                "sv_sva_grupp": "SV",
                "meritvarde_16": 160,
                "meritvarde_17": 170,
                "uppnatt_alla_amnen": True,
                "behorig_yrkesprogram": True,
                "behorig_hogskoleforberedande_nagot_program": False,
            },
            {
                "Skolenhetskod": "123",
                "kon": "Flickor",
                "sv_sva_grupp": "SVA",
                "meritvarde_16": 180,
                "meritvarde_17": 190,
                "uppnatt_alla_amnen": False,
                "behorig_yrkesprogram": False,
                "behorig_hogskoleforberedande_nagot_program": False,
            },
        ]
        result = overview(rows, "2025-2026", 9, {"123": "Testskolan"})

        segments = {(row["niva"], row["kon"], row["elevgrupp"]) for row in result}
        self.assertIn(("skolenhet", "Alla", "Alla"), segments)
        self.assertIn(("skolenhet", "Pojkar", "Alla"), segments)
        self.assertIn(("skolenhet", "Flickor", "Alla"), segments)
        self.assertIn(("skolenhet", "Alla", "SV"), segments)
        self.assertIn(("skolenhet", "Alla", "SVA"), segments)

    def test_overview_uses_kolada_rounding_for_grade6_all_subjects(self) -> None:
        rows = []
        for index in range(40):
          rows.append({
              "Skolenhetskod": "123",
              "kon": "Pojkar" if index % 2 else "Flickor",
              "sv_sva_grupp": "SV",
              "meritvarde_16": 160,
              "meritvarde_17": 170,
              "uppnatt_alla_amnen": index < 39,
          })

        result = overview(rows, "2025-2026", 6, {"123": "Testskolan"})
        total = next(row for row in result if row["niva"] == "alla_skolenheter" and row["kon"] == "Alla" and row["elevgrupp"] == "Alla")
        self.assertEqual(total["andel_uppnatt_alla_amnen"], 95.0)

    def test_control_rows_counts_valid_and_special_codes(self) -> None:
        rows = [
            {"Skolenhetskod": "123", "kon": "Pojkar", "sv_sva_grupp": "SV", "Sv": "A", "Ma": "F"},
            {"Skolenhetskod": "123", "kon": "Flickor", "sv_sva_grupp": "SV", "Sv": "2", "Ma": ""},
            {"Skolenhetskod": "123", "kon": "Flickor", "sv_sva_grupp": "SVA", "Sv": "", "Ma": "Y"},
        ]
        lookup = {"123": "Testskolan"}
        result = control_rows(rows, "2025-2026", 9, ["Sv", "Ma"], lookup)
        sv_all = next(row for row in result if row["niva"] == "alla_skolenheter" and row["elevgrupp"] == "Alla" and row["amne"] == "Sv")
        self.assertEqual(sv_all["antal_giltiga_betyg"], 1)
        self.assertEqual(sv_all["specialkod_2"], 1)
        self.assertEqual(sv_all["antal_tomma"], 1)
        ma_sv = next(row for row in result if row["niva"] == "skolenhet" and row["elevgrupp"] == "SV" and row["amne"] == "Ma")
        self.assertEqual(ma_sv["antal_F"], 1)
        self.assertEqual(ma_sv["antal_tomma"], 1)


class NpAggregationTests(unittest.TestCase):
    def test_aggregate_np_includes_gender_and_sv_sva_segments(self) -> None:
        np_rows_by_grade = [
            (
                {"Skolenhetskod": "123", "SV_KURSPLAN": "1", "SV_PROVB": "C", "EN_PROVB": "B", "MA_PROVB": "A"},
                {"kon": "Pojkar", "sv_sva_grupp": "SV", "Sv": "B", "Sva": "", "En": "B", "Ma": "A"},
            ),
            (
                {"Skolenhetskod": "123", "SV_KURSPLAN": "2", "SV_PROVB": "E", "EN_PROVB": "D", "MA_PROVB": "F"},
                {"kon": "Flickor", "sv_sva_grupp": "SVA", "Sv": "", "Sva": "E", "En": "C", "Ma": "F"},
            ),
        ]

        np_pass, np_relation = aggregate_np(np_rows_by_grade, "2025-2026", 6, {"123": "Testskolan"})

        boys_sv = next(
            row for row in np_pass
            if row["niva"] == "kommun" and row["kon"] == "Pojkar" and row["elevgrupp"] == "SV" and row["amne"] == "Sv/Sva"
        )
        self.assertEqual(boys_sv["antal_np"], 1)
        self.assertEqual(boys_sv["antal_godkanda_np"], 1)

        girls_sva = next(
            row for row in np_pass
            if row["niva"] == "kommun" and row["kon"] == "Flickor" and row["elevgrupp"] == "SVA" and row["amne"] == "Sv/Sva"
        )
        self.assertEqual(girls_sva["antal_np"], 1)

        girls_relation = next(
            row for row in np_relation
            if row["niva"] == "kommun" and row["kon"] == "Flickor" and row["elevgrupp"] == "SVA" and row["amne"] == "Sv/Sva"
        )
        self.assertEqual(girls_relation["antal_jamforda"], 1)
        self.assertEqual(girls_relation["antal_betyg_lika_np"], 1)

    def test_np_import_diagnostics_counts_documented_special_codes(self) -> None:
        rows = [
            {"MA_PROVB": "99", "SV_PROVB": "88", "EN_PROVB": "77"},
            {"MA_PROVB": "A", "SV_PROVB": "99", "EN_PROVB": ""},
        ]
        result = np_import_diagnostics(rows, [], NP_SPECS[6], "2025-2026", matched=1)

        self.assertEqual(result["special_codes_by_column"]["MA_PROVB"]["99"], 1)
        self.assertEqual(result["special_codes_by_column"]["SV_PROVB"]["88"], 1)
        self.assertEqual(result["special_codes_by_column"]["SV_PROVB"]["99"], 1)
        self.assertEqual(result["special_codes_by_column"]["EN_PROVB"]["77"], 1)


class ConfigPathsTests(unittest.TestCase):
    def test_resolve_paths_uses_default_project_year(self) -> None:
        paths = resolve_paths()
        self.assertEqual(paths.lasar, DEFAULT_LASAR)
        self.assertEqual(paths.output_dir, Path(paths.base_dir) / "data" / "output" / DEFAULT_LASAR)
        self.assertEqual(paths.json_dir, paths.output_dir / "json")

    def test_resolve_paths_allows_environment_override(self) -> None:
        with patch.dict("os.environ", {"BETYGSSTATISTIK_LASAR": "2024-2025"}, clear=False):
            paths = resolve_paths()
        self.assertEqual(paths.lasar, "2024-2025")
        self.assertEqual(paths.raw_betyg_dir, paths.raw_dir / "betyg" / "2024-2025")


if __name__ == "__main__":
    unittest.main()
